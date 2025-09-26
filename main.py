import openpyxl
import sys
import fonte
import time
import openpyxl.workbook
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bs4 import BeautifulSoup

inicio = time.time()


book = openpyxl.Workbook()
pagina = book['Sheet']
cabecalhos = ['Nome do livro', 'Preço em Euro', 'Preco em Real']


listaDeLivros = []
tamWidth = 0


cotacaoEur = 0

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'}



try:
    response = requests.get("https://economia.awesomeapi.com.br/json/last/EUR-BRL")
    response.raise_for_status()
    data = response.json()
    cotacaoEur = data['EURBRL']['bid']
    
except Exception as e:
    print(e)



try:
    for i in range(1,51): 
        response = requests.get(f'https://books.toscrape.com/catalogue/page-{i}.html', headers)

        soup = BeautifulSoup(response.content, 'html.parser')

        ol = soup.find_all('ol', {'class': 'row'})

        for i in ol:
            li = i.find_all('li')
            for l in li:
                h3 = l.find('h3')
                a = h3.find('a')
                text = a['title']

                divPrice = l.find('div', {'class': 'product_price'})
                p = divPrice.find('p')
                price = p.get_text()

                preco = price[1::]
                precoReal = float(preco) * float(cotacaoEur)
                precoFormatado = f"R$ {precoReal:.2f}"
                listaDeLivros.append((text, price, precoFormatado))

except Exception as e:
    print(e)

 
listaDeLivros.sort(key=lambda livro: livro[1], reverse=True)

try:
    pagina.append(cabecalhos)
    for texto, preco, precoReal in listaDeLivros:
        pagina.append([texto, preco, precoReal])


    for rows in pagina.iter_cols(max_row=1):
        for celula in rows:
            celula.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            celula.font = fonte.fonteTitulo

    for columns in pagina.iter_cols(min_col=1, max_col=3):
        tamWidth = 0
        for column in columns:         
            letraColuna = column.column_letter
            if column.value is not None:
                text = str(column.value)
                if (len(text) > tamWidth):
                    tamWidth = len(text)

        pagina.column_dimensions[letraColuna].width = tamWidth + 2

    for rows in pagina.iter_rows(min_row=1, max_row=pagina.max_row):
        for celula in rows:
            celula.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))


    book.save('Planilha de Livros.xlsx')

except Exception as e:
    print(e)

fim = time.time()

duracao = fim - inicio

print(f"Execução finalizada em {duracao:2f} Segundos")